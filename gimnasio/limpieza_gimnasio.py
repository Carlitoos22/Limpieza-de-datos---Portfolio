import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────
df = pd.read_excel('ejercicio_gimnasio_SUCIO.xlsx', dtype=str)

# ─────────────────────────────────────────────
# NIVEL 1 – Limpieza base
# ─────────────────────────────────────────────

# Eliminar filas completamente vacías
df.dropna(how='all', inplace=True)

# Eliminar filas donde Socio es basura (solo espacios, guiones, etc.)
df = df[df['Socio'].str.strip().str.len() > 2]

# Normalizar nombres: Title Case
df['Socio'] = df['Socio'].str.strip().str.title()

# Normalizar actividades
actividad_map = {
    'musculacion': 'Musculación',
    'musculación': 'Musculación',
    'pilates': 'Pilates',
    'spinning': 'Spinning',
    'crossfit': 'CrossFit',
    'yoga': 'Yoga',
}
df['Actividad'] = df['Actividad'].str.strip().str.lower().map(
    lambda x: actividad_map.get(x, x.title()) if isinstance(x, str) else x
)

# Limpiar precios → numérico
def limpiar_monto(val):
    if pd.isna(val):
        return np.nan
    val = str(val).strip().upper()
    if val in ('PENDIENTE', 'GRATIS', ''):
        return np.nan
    val = val.replace('$', '').replace(',', '').replace(' ', '')
    try:
        return float(val)
    except ValueError:
        return np.nan

df['Monto'] = df['Monto'].apply(limpiar_monto)

# Limpiar fechas
def limpiar_fecha(val):
    if pd.isna(val):
        return pd.NaT
    val = str(val).strip()
    # Normalizar formato corto 25 → 2025
    val = re.sub(r'(\d{2})/(\d{2})/(\d{2})$', r'\1/\2/20\3', val)
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return pd.to_datetime(val, format=fmt)
        except Exception:
            pass
    try:
        return pd.to_datetime(val, dayfirst=True)
    except Exception:
        return pd.NaT

df['Fecha Pago'] = df['Fecha Pago'].apply(limpiar_fecha)

# ─────────────────────────────────────────────
# NIVEL 2 – Limpieza avanzada
# ─────────────────────────────────────────────

# Emails a minúscula y reemplazar N/A, "-" por vacío
df['Email'] = df['Email'].str.strip().str.lower()
df['Email'] = df['Email'].replace({'n/a': np.nan, '-': np.nan, '': np.nan})

# Normalizar estados
estado_map = {
    'activo': 'Activo',
    'inactivo': 'Inactivo',
    'suspendido': 'Suspendido',
}
df['Estado'] = df['Estado'].str.strip().str.lower().map(
    lambda x: estado_map.get(x, x.title()) if isinstance(x, str) else x
)

# Normalizar métodos de pago
def normalizar_metodo(val):
    if pd.isna(val):
        return np.nan
    val = str(val).strip().lower()
    if val in ('mp', 'mercadopago', 'mercado pago'):
        return 'Mercado Pago'
    if val in ('transferencia', 'transf', 'transfer'):
        return 'Transferencia'
    if val in ('tarjeta debito', 'tarjeta débito', 'tarjeta', 'tarjeta crédito', 'tarjeta credito'):
        return 'Tarjeta'
    if val == 'efectivo':
        return 'Efectivo'
    return val.title()

df['Metodo Pago'] = df['Metodo Pago'].apply(normalizar_metodo)

# Eliminar duplicados exactos
df.drop_duplicates(inplace=True)
df.reset_index(drop=True, inplace=True)

# ─────────────────────────────────────────────
# NIVEL 3 – Reportes para Diego
# ─────────────────────────────────────────────

total_facturado = df['Monto'].sum()

conteo_estado = df['Estado'].value_counts()
activos = conteo_estado.get('Activo', 0)
inactivos = conteo_estado.get('Inactivo', 0)
suspendidos = conteo_estado.get('Suspendido', 0)

fact_actividad = (
    df.groupby('Actividad')['Monto']
    .sum()
    .sort_values(ascending=False)
    .reset_index()
)
fact_actividad.columns = ['Actividad', 'Total Facturado']

metodo_frecuencia = (
    df['Metodo Pago']
    .value_counts()
    .reset_index()
)
metodo_frecuencia.columns = ['Método de Pago', 'Cantidad']

# ─────────────────────────────────────────────
# EXPORTAR A EXCEL
# ─────────────────────────────────────────────

wb = Workbook()

# ── Estilos ──────────────────────────────────
AZUL_HEADER = "1F3864"
AZUL_CLARO  = "D6E4F0"
VERDE       = "1B5E20"
VERDE_CLARO = "E8F5E9"
GRIS        = "F5F5F5"
BLANCO      = "FFFFFF"

def header_style(cell, bg=AZUL_HEADER, bold=True, font_color="FFFFFF"):
    cell.font = Font(bold=bold, color=font_color, name='Arial', size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def thin_border():
    s = Side(style='thin', color='BDBDBD')
    return Border(left=s, right=s, top=s, bottom=s)

def apply_borders(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ══════════════════════════════════════════════
# HOJA 1 – Datos Limpios
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Datos Limpios"
ws1.freeze_panes = "A2"

cols = ['Socio', 'Email', 'Telefono', 'Actividad', 'Monto', 'Fecha Pago', 'Estado', 'Metodo Pago']
ws1.append(cols)

for cell in ws1[1]:
    header_style(cell)
ws1.row_dimensions[1].height = 28

for i, row in df[cols].iterrows():
    ws1.append([
        row['Socio'],
        row['Email'] if pd.notna(row['Email']) else '',
        row['Telefono'] if pd.notna(row['Telefono']) else '',
        row['Actividad'] if pd.notna(row['Actividad']) else '',
        row['Monto'] if pd.notna(row['Monto']) else '',
        row['Fecha Pago'].strftime('%d/%m/%Y') if pd.notna(row['Fecha Pago']) else '',
        row['Estado'] if pd.notna(row['Estado']) else '',
        row['Metodo Pago'] if pd.notna(row['Metodo Pago']) else '',
    ])
    excel_row = i + 2
    fill = PatternFill("solid", start_color=GRIS if i % 2 == 0 else BLANCO)
    for cell in ws1[excel_row]:
        cell.fill = fill
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border()
    # Monto numérico formateado
    monto_cell = ws1.cell(row=excel_row, column=5)
    if row['Monto'] and pd.notna(row['Monto']):
        monto_cell.number_format = '$#,##0.00'

set_col_widths(ws1, {
    'A': 22, 'B': 28, 'C': 16, 'D': 14,
    'E': 13, 'F': 14, 'G': 13, 'H': 16
})

# ══════════════════════════════════════════════
# HOJA 2 – Resumen Ejecutivo
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("Resumen Ejecutivo")

# Título principal
ws2.merge_cells('A1:F1')
ws2['A1'] = "📊  REPORTE DE GESTIÓN – GIMNASIO DIEGO"
ws2['A1'].font = Font(bold=True, size=16, color=BLANCO, name='Arial')
ws2['A1'].fill = PatternFill("solid", start_color=AZUL_HEADER)
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 38

# ── Bloque KPIs ──────────────────────────────
ws2['A3'] = "INDICADORES CLAVE"
ws2['A3'].font = Font(bold=True, size=11, color=AZUL_HEADER, name='Arial')

kpis = [
    ("💰 Total Facturado", f"${total_facturado:,.2f}"),
    ("✅ Socios Activos", activos),
    ("❌ Socios Inactivos", inactivos),
    ("⏸  Socios Suspendidos", suspendidos),
    ("📋 Total Registros", len(df)),
]

for r, (label, value) in enumerate(kpis, start=4):
    ws2.cell(row=r, column=1, value=label).font = Font(name='Arial', size=10, bold=True)
    val_cell = ws2.cell(row=r, column=2, value=value)
    val_cell.font = Font(name='Arial', size=10, color=VERDE)
    val_cell.alignment = Alignment(horizontal='center')
    for c in [1, 2]:
        cell = ws2.cell(row=r, column=c)
        cell.fill = PatternFill("solid", start_color=VERDE_CLARO if r % 2 == 0 else BLANCO)
        cell.border = thin_border()

# ── Facturación por Actividad ─────────────────
start_act = 4
ws2['D3'] = "FACTURACIÓN POR ACTIVIDAD"
ws2['D3'].font = Font(bold=True, size=11, color=AZUL_HEADER, name='Arial')

for cell in ws2[start_act]:
    pass  # placeholder

header_cells = [ws2.cell(row=start_act, column=4), ws2.cell(row=start_act, column=5)]
ws2.cell(row=start_act, column=4, value="Actividad")
ws2.cell(row=start_act, column=5, value="Total ($)")
for c in [4, 5]:
    header_style(ws2.cell(row=start_act, column=c))

for i, row in fact_actividad.iterrows():
    r = start_act + 1 + i
    ws2.cell(row=r, column=4, value=row['Actividad']).font = Font(name='Arial', size=10)
    amt = ws2.cell(row=r, column=5, value=row['Total Facturado'])
    amt.font = Font(name='Arial', size=10)
    amt.number_format = '$#,##0.00'
    bg = AZUL_CLARO if i % 2 == 0 else BLANCO
    for c in [4, 5]:
        ws2.cell(row=r, column=c).fill = PatternFill("solid", start_color=bg)
        ws2.cell(row=r, column=c).border = thin_border()

# ── Método de pago ────────────────────────────
start_mp = start_act + len(fact_actividad) + 3
ws2.cell(row=start_mp - 1, column=4, value="MÉTODO DE PAGO MÁS USADO").font = Font(
    bold=True, size=11, color=AZUL_HEADER, name='Arial'
)

ws2.cell(row=start_mp, column=4, value="Método")
ws2.cell(row=start_mp, column=5, value="Cantidad")
for c in [4, 5]:
    header_style(ws2.cell(row=start_mp, column=c))

for i, row in metodo_frecuencia.iterrows():
    r = start_mp + 1 + i
    ws2.cell(row=r, column=4, value=row['Método de Pago']).font = Font(name='Arial', size=10)
    ws2.cell(row=r, column=5, value=row['Cantidad']).font = Font(name='Arial', size=10)
    bg = AZUL_CLARO if i % 2 == 0 else BLANCO
    for c in [4, 5]:
        ws2.cell(row=r, column=c).fill = PatternFill("solid", start_color=bg)
        ws2.cell(row=r, column=c).border = thin_border()

set_col_widths(ws2, {'A': 26, 'B': 18, 'C': 3, 'D': 22, 'E': 16, 'F': 3})
ws2.row_dimensions[3].height = 20

# ══════════════════════════════════════════════
# GUARDAR
# ══════════════════════════════════════════════
output_path = 'gimnasio_LIMPIO.xlsx'
wb.save(output_path)
print(f"✅ Archivo guardado: {output_path}")
print(f"   Registros limpios: {len(df)}")
print(f"   Total facturado:   ${total_facturado:,.2f}")
print(f"   Activos / Inactivos / Suspendidos: {activos} / {inactivos} / {suspendidos}")
print(f"   Actividad top: {fact_actividad.iloc[0]['Actividad']} (${fact_actividad.iloc[0]['Total Facturado']:,.2f})")
print(f"   Método más usado: {metodo_frecuencia.iloc[0]['Método de Pago']} ({metodo_frecuencia.iloc[0]['Cantidad']} usos)")
